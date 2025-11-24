from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
import os
from datetime import datetime
import sys
import traceback # 導入 traceback 以便詳細輸出錯誤

# 創建 Flask 應用程式實例
app = Flask(__name__)

# 啟用 CORS（跨來源資源共享），這允許前端從不同的來源請求 API。
CORS(app)

# =========================================================
# 檔案和工作表設定
# =========================================================
# 獲取應用程式的根目錄，使其可以在打包成 .exe 後也能正常運作
def get_app_path():
    """判斷應用程式是從原始腳本還是從 .exe 運行，並返回正確的路徑。"""
    if getattr(sys, 'frozen', False):
        # 如果是從 PyInstaller 打包的 .exe 運行
        return os.path.dirname(sys.executable)
    else:
        # 如果是從原始 Python 腳本運行
        return os.path.dirname(os.path.abspath(__file__))

# 獲取應用程式的動態路徑
APP_ROOT_PATH = get_app_path()

# 設定 Excel 檔案路徑為應用程式根目錄下的 warehouse_data.xlsx
EXCEL_FILE = os.path.join(APP_ROOT_PATH, 'warehouse_data.xlsx')

MAP_SHEET_NAME = 'map'
OPTIONS_SHEET_NAME = 'rowdown'
DATA_SHEET_NAME = 'test'
# 客戶資料工作表名稱
CUSTOMER_SHEET_NAME = 'sale_sheet_data'
# 新增商品資料工作表名稱
GOODS_SHEET_NAME = 'goods_sheet'
# 銷售下拉式選單
SALE_ROWDOWN_NAME = 'rowdown_order'

# 檢查 Excel 檔案是否存在，如果不存在則創建一個新的並初始化工作表
if not os.path.exists(EXCEL_FILE):
    try:
        wb = openpyxl.Workbook()
        wb.create_sheet(MAP_SHEET_NAME, 0)
        wb.create_sheet(OPTIONS_SHEET_NAME, 1)
        wb.create_sheet(DATA_SHEET_NAME, 2)
        wb.create_sheet(CUSTOMER_SHEET_NAME, 3) # 新增客戶資料表
        wb.create_sheet(GOODS_SHEET_NAME, 4) # 新增商品資料表
        wb.create_sheet(SALE_ROWDOWN_NAME, 5)
        wb.remove(wb['Sheet'])

        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"創建 Excel 檔案時發生錯誤: {e}")

# =========================================================
# 前端靜態檔案伺服器
# =========================================================
# 新增一個根路由，用於提供前端網頁 (index.html)
@app.route('/')
@app.route('/index.html')
def serve_index():
    return send_from_directory(APP_ROOT_PATH, 'index.html')

# 新增 order 頁面路由，用於提供 order.html
@app.route('/order')
@app.route('/order.html')
def serve_order():
    return send_from_directory(APP_ROOT_PATH, 'order.html')

# 新增 sale 頁面路由，用於提供 sale.html
@app.route('/sale')
@app.route('/sale.html')
def serve_sale():
    return send_from_directory(APP_ROOT_PATH, 'sale.html')


# =========================================================
# API 路由 - 查詢客戶資料 (更新為多欄位查詢)
# =========================================================
@app.route('/api/customers', methods=['GET'])
def get_customers():
    """根據多個欄位（編號、名稱、電話、地址）查詢客戶資料，使用精準前綴比對。"""
    try:
        # 1. 獲取所有查詢參數，並轉換為小寫以便進行前綴比對
        id_query = request.args.get('id_query', '').strip().lower()
        name_query = request.args.get('name_query', '').strip().lower()
        phone_query = request.args.get('phone_query', '').strip().lower()
        address_query = request.args.get('address_query', '').strip().lower()

        # 檢查是否所有查詢字串都為空，如果是，則返回空列表
        if not (id_query or name_query or phone_query or address_query):
            return jsonify([])

        # 注意：此處仍保留每次請求都讀取 Excel 檔案的邏輯，會影響效能。
        # 建議參考先前提供的 customer_search_backend.py 檔案，使用記憶體快取。
        wb = openpyxl.load_workbook(EXCEL_FILE)
        customer_sheet = wb[CUSTOMER_SHEET_NAME]
        
        customers = []
        
        # 獲取標題列 (假設是第一行)
        headers = [str(cell.value) for cell in customer_sheet[1]]
        
        # 確定欄位索引，用於查詢和映射
        header_map = {
            '客戶編號': -1, '客戶名稱': -1, '客戶電話': -1, '送貨地址': -1
        }
        for key in header_map.keys():
            try:
                header_map[key] = headers.index(key)
            except ValueError:
                # 即使缺少欄位也不中斷，但會記錄
                print(f"Warning: Missing header '{key}' in sheet '{CUSTOMER_SHEET_NAME}'")

        # 從第二行開始遍歷數據
        for row in customer_sheet.iter_rows(min_row=2):
            # 將行數據轉換為字典
            row_data = [str(cell.value or '').strip() for cell in row]
            
            customer_info = {}
            for i, header in enumerate(headers):
                # 如果行數據長度不足，使用空字符串
                customer_info[header] = row_data[i] if i < len(row_data) else ''

            is_match = False
            
            # --- 核心修改：將 'in' 替換為 '.startswith()' 實現前綴比對 ---
            
            # 1. 客戶編號比對
            id_val = customer_info.get('客戶編號', '').lower()
            if id_query and id_val and id_val.startswith(id_query):
                is_match = True
            
            # 2. 客戶名稱比對
            name_val = customer_info.get('客戶名稱', '').lower()
            if not is_match and name_query and name_val and name_val.startswith(name_query):
                is_match = True

            # 3. 客戶電話比對
            phone_val = customer_info.get('客戶電話', '').lower()
            if not is_match and phone_query and phone_val and phone_val.startswith(phone_query):
                is_match = True

            # 4. 送貨地址比對
            address_val = customer_info.get('送貨地址', '').lower()
            if not is_match and address_query and address_val and address_val.startswith(address_query):
                is_match = True

            if is_match:
                # 映射 Excel 欄位（中文）到前端需要的英文字段
                formatted_customer = {
                    'id': customer_info.get('客戶編號', ''),
                    'name': customer_info.get('客戶名稱', ''),
                    'phone': customer_info.get('客戶電話', ''),
                    'address': customer_info.get('送貨地址', '')
                }
                customers.append(formatted_customer)
        
        return jsonify(customers)
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

# =========================================================
# API 路由 - 查詢商品資料 (新增)
# =========================================================
@app.route('/api/goods', methods=['GET'])
def get_goods_data():
    """從 goods_sheet 工作表讀取所有商品資料。"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        goods_sheet = wb[GOODS_SHEET_NAME]

        goods_data = []
        # 獲取標題列並過濾掉 None 值
        headers = [str(cell.value).strip() for cell in goods_sheet[1] if cell.value]

        # 從第二行開始遍歷數據
        for row in goods_sheet.iter_rows(min_row=2):
            # 將行數據轉換為字串
            # row_data = [str(cell.value or '').strip() for cell in row]
            row_dict = {headers[i]: (str(cell.value).strip() if cell.value else '') 
                        for i, cell in enumerate(row)}
            
            # # 將行數據映射到字典
            # item = {}
            # for i, header in enumerate(headers):
            #     if i < len(row_data):
            #         item[header] = row_data[i]
            
            # 為了前端方便，將中文 key 轉換為英文 key
            formatted_item = {
                'name': row_dict.get('品名', ''),
                'spec': row_dict.get('規格', ''),
                'stock': row_dict.get('庫存', ''),
            }
            goods_data.append(formatted_item)

        return jsonify(goods_data)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"讀取商品資料時發生錯誤: {str(e)}"}), 500
    
# --------------------------------------------------------------------------------
# API 路由 - 獲取銷售頁面下拉式選單資料 (GET)
# --------------------------------------------------------------------------------
@app.route('/api/rowdown_order', methods=['GET'])
def get_rowdown_data():
    # 將 Excel 標題與 JSON 鍵名的對應關係定義在函式內部
    HEADER_TO_JSON_KEY = {
        '計價單位': 'pricingUnits',
        '銷售方式': 'salesMethods',
        '製單人員': 'creatorNames',
        '送貨員': 'deliveryPeople',
        '車號': 'carNumbers',
    }  
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        if SALE_ROWDOWN_NAME not in wb.sheetnames:
            return jsonify({"status": "error", "message": f"Excel中找不到工作表: {SALE_ROWDOWN_NAME}"}), 500
            
        options_sheet = wb[SALE_ROWDOWN_NAME]
        options_values = list(options_sheet.values)
        
        lookup_data = {}

        if options_values:
            # 2. 提取標頭 (第一行) 和數據行 (從第二行開始)
            options_header = options_values[0]
            options_data_rows = options_values[1:]
            
            # 3. 遍歷每個欄位，提取數據並進行鍵名轉換
            for col_index, header in enumerate(options_header):
                # 檢查標頭是否在我們需要的對應列表中
                # 使用 str() 確保 header 是一個字串，以避免 NoneType 錯誤
                if isinstance(header, str) and header in HEADER_TO_JSON_KEY:
                    
                    # 獲取目標 JSON 鍵名
                    json_key = HEADER_TO_JSON_KEY[header]
                    options_list = []
                    seen_options = set() # 用於去重 (優化)
                    
                    # 遍歷所有數據行，提取該欄位的值
                    for row in options_data_rows:
                        # 檢查該行在當前欄位是否有值
                        if len(row) > col_index and row[col_index] is not None:
                            value = str(row[col_index]).strip()
                            
                            # 檢查值是否非空且尚未出現過
                            if value and value not in seen_options:
                                seen_options.add(value)
                                options_list.append(value)
                                
                    # 將結果存入字典，使用轉換後的 JSON 鍵名
                    lookup_data[json_key] = options_list

        # 4. 返回 JSON 格式的下拉選單數據
        return jsonify({
            "lookupData": lookup_data
        })
        
    except FileNotFoundError:
        return jsonify({"status": "error", "message": f"找不到 Excel 檔案: {EXCEL_FILE}"}), 500
    except Exception as e:
        traceback.print_exc()# 輸出錯誤堆棧，方便除錯
        return jsonify({"status": "error", "message": f"處理資料時發生錯誤: {str(e)}"}), 500


# =========================================================
# API 路由 - 處理 GET 請求 (保留原有功能)
# =========================================================
@app.route('/api/data', methods=['GET'])
def get_data():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        options_sheet = wb[OPTIONS_SHEET_NAME]
        options_values = list(options_sheet.values)
        dropdown_options = {}
        if options_values:
            options_header = options_values[0]
            options_data_rows = options_values[1:]
            
            for col_index, header in enumerate(options_header):
                if header:
                    options_list = []
                    seen_options = set()
                    for row in options_data_rows:
                        if len(row) > col_index and row[col_index] is not None:
                            value = str(row[col_index]).strip()
                            if value and value not in seen_options:
                                seen_options.add(value)
                                options_list.append(value)
                    dropdown_options[header] = options_list

        map_sheet = wb[MAP_SHEET_NAME]
        bin_map_data = []
        for row in range(1, map_sheet.max_row + 1):
            for col in range(1, map_sheet.max_column + 1):
                cell_value = map_sheet.cell(row, col).value
                position_name = f"{col}-{row}"
                
                if cell_value and isinstance(cell_value, str):
                    lines = cell_value.split('\n')
                    bin_map_data.append({
                        'positionName': position_name,
                        'binName': lines[0] if len(lines) > 0 else '',
                        'item': lines[1] if len(lines) > 1 else '',
                        'date': lines[2] if len(lines) > 2 else '',
                        'vendor': lines[3] if len(lines) > 3 else '',
                        'binValue': cell_value
                    })
                else:
                    bin_map_data.append({
                        'positionName': position_name,
                        'binName': '',
                        'binValue': ''
                    })
        
        return jsonify({
            "dropdownOptions": dropdown_options,
            "binMapData": bin_map_data
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

# =========================================================
# API 路由 - 處理 POST 請求 (保留原有功能)
# =========================================================
@app.route('/api/submit', methods=['POST'])
def submit_data():
    try:
        form_data = request.form
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        timestamp = datetime.now()
        
        data_sheet = wb[DATA_SHEET_NAME]
        new_row = [
            timestamp,
            form_data.get('日期'),
            form_data.get('輸入原物料'),
            form_data.get('廠商名稱'),
            form_data.get('乾燥度'),
            form_data.get('等級'),
            form_data.get('料桶'),
            form_data.get('總重'),
            form_data.get('容量'),
            form_data.get('初估碾米率'),
            form_data.get('備註')
        ]
        data_sheet.append(new_row)
        
        map_sheet = wb[MAP_SHEET_NAME]
        bin_position = form_data.get('positionName')
        bin_display_name = form_data.get('料桶')
        
        if not bin_position:
            raise ValueError("缺少 'positionName' 參數")
            
        col_index, row_index = map(int, bin_position.split('-'))
        
        new_cell_value = (
            f"{bin_display_name}\n"
            f"{form_data.get('輸入原物料')}\n"
            f"{form_data.get('日期')}\n"
            f"{form_data.get('廠商名稱')}"
        )
        map_sheet.cell(row=row_index, column=col_index, value=new_cell_value)
        
        wb.save(EXCEL_FILE)
        return jsonify({"status": "success"})
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

# =========================================================
# 主程式入口點
# =========================================================
if __name__ == '__main__':
    # 確保 Flask 在您指定的端口運行
    app.run(debug=True, host='0.0.0.0', port=5719)
